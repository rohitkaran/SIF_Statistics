#!/usr/bin/env python3
"""
fetch_sif_portfolios.py  --  Download & normalise fund-house PORTFOLIO disclosures for SIFs.

SIFs (Specialized Investment Funds) disclose their holdings BI-MONTHLY (every alternate
month -- a SEBI rule), NOT monthly. AMFI does not host the holdings: its "Portfolio
Disclosure" page, like the mutual-fund one, only links out to each AMC's own website. So --
unlike NAVs (fetch_amfi_sif.py, which hits real AMFI JSON APIs) -- portfolios must be pulled
per-AMC, from heterogeneous Excel/PDF files on ~17 AMC sites.

This script is registry-driven (portfolios/registry.json). Each AMC entry has:
  * discovery -- HOW to find the file for a period:
      http_listing : regex a listing page for the file URL (+ date)
      template     : build candidate URLs from month-end dates via a strftime template
      manual       : pinned {period, as_of, url} list (for WAF/SPA/API-only sites)
  * adapter   -- HOW to parse it (currently sebi_xlsx: the SEBI-standard monthly-portfolio
                 layout every AMC uses; also unpacks a zip-of-xlsx bundle).

All output is COMMITTED to the repo: raw files verbatim under portfolios/raw/, normalised
holdings under portfolios/data/, and portfolios/index.json mapping each dashboard scheme code
to the periods available. Adding an AMC = one registry entry (+ a new adapter only if the file
layout is genuinely different).

Requires: openpyxl (xlsx), pdfplumber (pdf). Install: pip install -r requirements.txt

Examples:
  python fetch_sif_portfolios.py                       # all AMCs, latest disclosed period
  python fetch_sif_portfolios.py --amc mirae
  python fetch_sif_portfolios.py --file mine.xlsx --amc apex --period 2026-05
  python fetch_sif_portfolios.py --list
"""

import argparse
import calendar
import datetime as dt
import io
import json
import os
import re
import ssl
import statistics
import sys
import urllib.error
import urllib.parse
import urllib.request
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

_SSL_UNVERIFIED = ssl.create_default_context()
_SSL_UNVERIFIED.check_hostname = False
_SSL_UNVERIFIED.verify_mode = ssl.CERT_NONE

ISIN_RE = re.compile(r"^IN[A-Z0-9]{10}$")


# --------------------------------------------------------------------------- #
# HTTP  (mirrors the NAV fetcher: browser UA, retries, SSL fallback)
# --------------------------------------------------------------------------- #

def http_get(url, timeout=60, retries=3, backoff=1.6):
    url = urllib.parse.quote(url, safe=":/?#[]@!$&'()*+,;=%~")   # encode stray spaces etc.
    headers = {"User-Agent": UA, "Accept": "*/*", "Accept-Encoding": "identity",
               "Accept-Language": "en-US,en;q=0.9"}
    last = None
    for attempt in range(retries):
        for ctx in (ssl.create_default_context(), _SSL_UNVERIFIED):
            try:
                req = urllib.request.Request(url, headers=headers)
                with urllib.request.urlopen(req, timeout=timeout, context=ctx) as r:
                    return r.read()
            except urllib.error.HTTPError as e:
                if 400 <= e.code < 500:
                    raise
                last = e
            except ssl.SSLError as e:
                last = e
                continue
            except (urllib.error.URLError, TimeoutError, ConnectionError) as e:
                last = e
        import time
        time.sleep(backoff ** attempt)
    raise last if last else RuntimeError("unreachable")


def http_ok(url, timeout=30):
    """True if the URL returns 200 with real spreadsheet/zip content (not a WAF/HTML page)."""
    try:
        blob = http_get(url, timeout=timeout)
    except Exception:
        return None
    if looks_like_office(blob) or blob[:4] == b"%PDF":
        return blob
    return None


# --------------------------------------------------------------------------- #
# small helpers
# --------------------------------------------------------------------------- #

def norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("%", "").replace("$", "")
    if not s or s in ("-", "NA", "N.A.", "nan"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def looks_like_office(blob):
    """xlsx/zip are ZIP (PK); .xls is OLE (D0CF11E0)."""
    return blob[:2] == b"PK" or blob[:4] == b"\xd0\xcf\x11\xe0"


def load_registry():
    with open(os.path.join(HERE, "portfolios", "registry.json"), encoding="utf-8") as f:
        return json.load(f)


def load_nav_schemes():
    path = os.path.join(HERE, "nav_data.json")
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        d = json.load(f)
    return [{"code": s["code"], "name": s.get("name", ""), "sif": s.get("sif", "")}
            for s in d.get("schemes", [])]


def strat_token(text):
    """Classify a fund/scheme name into its SIF strategy family (order matters)."""
    t = (text or "").lower()
    if "ex" in t and "top" in t and "100" in t:
        return "extop"
    if "sector" in t and "rotation" in t:
        return "sector"
    if "allocator" in t or "asset alloc" in t:
        return "alloc"
    if "hybrid" in t:
        return "hybrid"
    if "equity" in t and "long" in t and "short" in t:
        return "equitylong"
    return None


def base_name(name):
    """'X Fund - Direct Plan - Growth' -> 'X Fund' for a clean display label."""
    return re.split(r"\s*[-–]\s*(?:Direct|Regular)", name)[0].strip()


def match_codes(fund_name, amc_schemes):
    fn = norm(fund_name)
    if not fn:
        return []
    core = fn
    for tail in ("fund", "sif", "directplan", "regularplan", "growth", "idcw"):
        core = core.replace(tail, "")
    out = []
    for s in amc_schemes:
        sn = norm(s["name"])
        if sn.startswith(fn) or fn in sn or (len(core) >= 8 and core in sn):
            out.append(s["code"])
    return out


def map_fund(ident, fund_name, amc_schemes):
    """(codes, display_name) for a disclosed fund, matched to dashboard schemes by strategy."""
    strat = strat_token(ident + " " + fund_name)
    codes, disp = [], fund_name
    if strat:
        codes = [s["code"] for s in amc_schemes if strat_token(s["name"]) == strat]
    if not codes:
        codes = match_codes(fund_name, amc_schemes)
    # if the AMC has exactly one strategy family, assign all of it
    if not codes and amc_schemes:
        fams = {strat_token(s["name"]) for s in amc_schemes if strat_token(s["name"])}
        if len(fams) == 1:
            codes = [s["code"] for s in amc_schemes]
    if codes:
        by = {s["code"]: s["name"] for s in amc_schemes}
        disp = base_name(by[codes[0]])
    return codes, strat, disp


def month_ends(n=10, today=None):
    """Last n month-end dates, newest first, starting from the most recent completed month."""
    today = today or dt.date.today()
    y, m = today.year, today.month
    # start at previous month end (current month usually not disclosed yet)
    out = []
    for _ in range(n):
        m -= 1
        if m == 0:
            m = 12
            y -= 1
        out.append(dt.date(y, m, calendar.monthrange(y, m)[1]))
    return out


# --------------------------------------------------------------------------- #
# discovery
# --------------------------------------------------------------------------- #

def discover_http_listing(cfg):
    """Regex listing page(s) for the file URL; date from a named group in file_pattern."""
    disc = cfg["discovery"]
    pat = re.compile(disc["file_pattern"])
    fmt = disc.get("date_format", "%d%m%Y")
    found = {}
    for listing in disc.get("listing_urls", [cfg["site"]]):
        try:
            html = http_get(listing).decode("utf-8", "replace")
        except Exception as e:
            sys.stderr.write(f"  ! {listing}: {e}\n")
            continue
        for m in pat.finditer(html):
            path = m.group(0)
            gd = m.groupdict()
            try:
                d = dt.datetime.strptime(gd["date"], fmt).date() if gd.get("date") else None
            except ValueError:
                d = None
            # resolve relative paths against the listing's ORIGIN (not its sub-path)
            url = path if path.startswith("http") else urllib.parse.urljoin(listing, path)
            per = f"{d.year}-{d.month:02d}" if d else "unknown"
            found[url] = {"period": per, "as_of": d.isoformat() if d else "",
                          "url": url, "filename": os.path.basename(urllib.parse.urlparse(url).path)}
    return sorted(found.values(), key=lambda x: x["as_of"])


def _tokens(d):
    up = d.replace(day=28) + dt.timedelta(days=7)   # a day in the following month
    return {
        "y": d.year, "yy": f"{d.year % 100:02d}", "m": f"{d.month:02d}", "dd": f"{d.day:02d}",
        "mon": d.strftime("%b").lower(), "Mon": d.strftime("%b"), "MON": d.strftime("%b").upper(),
        "month": d.strftime("%B").lower(), "Month": d.strftime("%B"),
        "uy": up.year, "um": f"{up.month:02d}",   # upload-month (data month + 1)
    }


def discover_template(cfg):
    """Probe candidate month-end URLs built from a strftime-ish template; keep the 200s."""
    disc = cfg["discovery"]
    tmpl = disc["url_template"]
    out = []
    for d in month_ends(disc.get("lookback", 8)):
        url = tmpl.format(**_tokens(d))
        blob = http_ok(url)
        if blob is not None:
            out.append({"period": f"{d.year}-{d.month:02d}", "as_of": d.isoformat(),
                        "url": url, "filename": os.path.basename(urllib.parse.urlparse(url).path)
                                     or f"{cfg.get('adapter','file')}_{d.isoformat()}.xlsx",
                        "_blob": blob})
    return sorted(out, key=lambda x: x["as_of"])


def discover_manual(cfg):
    """Pinned URLs in the registry (for WAF/SPA/API-only AMCs)."""
    out = []
    for f in cfg["discovery"].get("files", []):
        d = f.get("as_of", "")
        per = f.get("period") or (d[:7] if d else "unknown")
        out.append({"period": per, "as_of": d, "url": f["url"],
                    "filename": f.get("filename") or os.path.basename(urllib.parse.urlparse(f["url"]).path)})
    return sorted(out, key=lambda x: x["as_of"])


DISCOVERERS = {"http_listing": discover_http_listing, "template": discover_template,
               "manual": discover_manual}


# --------------------------------------------------------------------------- #
# SEBI-standard xlsx parser  (every AMC's monthly-portfolio layout)
# --------------------------------------------------------------------------- #

HEADER_KEYS = {
    "name": ("name of the instrument", "name of instrument", "instrument", "security name", "company"),
    "isin": ("isin",),
    "industry": ("industry", "rating", "sector"),
    "qty": ("quantity", "no. of", "no of", "units"),
    "mv": ("market value", "market/fair value", "fair value", "exposure/market value",
           "market value(rs", "market value (rs"),
    "pct": ("% to net asset", "% to nav", "% to net assets", "percentage to net asset",
            "% of net asset", "% to aum", "% of aum"),
}


def _ws(s):
    """lower-case + collapse runs of whitespace (AMC headers vary: 'Net  Assets')."""
    return re.sub(r"\s+", " ", str(s).lower()).strip() if s is not None else ""


def _find_header(rows):
    for i, row in enumerate(rows):
        cells = [_ws(c) for c in row]
        joined = " | ".join(cells)
        has_name = any(k in joined for k in HEADER_KEYS["name"])
        has_pct = any(k in joined for k in HEADER_KEYS["pct"])
        if has_name and has_pct:
            def col(keys):
                for k in keys:
                    for j, c in enumerate(cells):
                        if k in c:
                            return j
                return -1
            return i, {k: col(v) for k, v in HEADER_KEYS.items()}
    return -1, None


def _classify(label):
    l = (label or "").lower()
    if any(k in l for k in ("money market", "treasury", "t-bill", "tbill", "cblo", "treps", "repo")):
        return "money_market"
    if any(k in l for k in ("debt", "bond", "debenture", "ncd", "g-sec", "government sec", "gilt")):
        return "debt"
    if "equity" in l:
        return "equity"
    return None


_MONS = {m.lower(): i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], 1)}


def _extract_date(text):
    """Find a portfolio 'as on' date in free text -> ISO 'YYYY-MM-DD', else None."""
    t = str(text)
    m = re.search(r"(\d{1,2})[-/\s]([A-Za-z]{3,9})[-/\s](\d{4})", t)
    if m and m.group(2)[:3].lower() in _MONS:
        return f"{int(m.group(3)):04d}-{_MONS[m.group(2)[:3].lower()]:02d}-{int(m.group(1)):02d}"
    m = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})", t)
    if m:
        return f"{int(m.group(3)):04d}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    return None


def parse_sebi_xlsx_one(ws, meta, amc, source_url, raw_rel):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hi, c = _find_header(rows)
    if hi < 0:
        return None
    # fund name + statement date + identity text from the rows above the header
    fund_name = ws.title
    as_of = meta["as_of"]
    ident_parts = [ws.title, meta.get("filename", "")]
    for r in rows[:hi]:
        for cell in r:
            if not cell:
                continue
            s = str(cell).strip()
            ident_parts.append(s)
            low = s.lower()
            # prefer a line that actually names the long-short fund
            if ("long" in low and "short" in low and "fund" in low
                    and "mutual fund" not in low and not _extract_date(s)):
                fund_name = s
            if ("as on" in low or "as at" in low or "statement" in low):
                d = _extract_date(s)
                if d:
                    as_of = d
    ident = " ".join(ident_parts)
    period = as_of[:7] if as_of else meta["period"]

    holdings, sectors, derivatives = [], {}, []
    cur_cat = None
    for r in rows[hi + 1:]:
        def cell(k):
            j = c[k]
            return r[j] if 0 <= j < len(r) else None
        name = (str(cell("name")).strip() if cell("name") is not None else "")
        isin = (str(cell("isin")).strip() if cell("isin") is not None else "")
        pct = to_float(cell("pct"))
        low = name.lower()
        if low == "grand total":
            break
        cat = _classify(name)
        if name and not ISIN_RE.match(isin):
            if cat:
                cur_cat = cat
            continue
        if ISIN_RE.match(isin) and pct is not None:
            mv = to_float(cell("mv"))
            hcat = cur_cat or "equity"
            h = {"name": name, "isin": isin,
                 "sector": (str(cell("industry")).strip() if cell("industry") else ""),
                 "qty": to_float(cell("qty")),
                 "mv_cr": round(mv / 100, 4) if mv is not None else None,   # Rs.Lakh -> Cr
                 "weight": round(pct * 100, 4), "category": hcat, "position": "Long"}
            holdings.append(h)
            if hcat == "equity" and h["sector"]:
                sectors[h["sector"]] = round(sectors.get(h["sector"], 0) + h["weight"], 4)

    derivatives = _derivatives(rows)
    if not holdings:
        return None

    net = None
    ratios = [h["mv_cr"] / (h["weight"] / 100) for h in holdings
              if h.get("mv_cr") and h["weight"] and h["weight"] > 0.05]
    if ratios:
        net = round(statistics.median(ratios), 2)
    alloc = {}
    for k in ("equity", "debt", "money_market"):
        s = round(sum(h["weight"] for h in holdings if h["category"] == k), 2)
        if s:
            alloc[k] = s
    alloc["cash_other"] = round(max(0.0, 100 - sum(alloc.values())), 2)

    holdings.sort(key=lambda h: -(h["weight"] or 0))
    return {
        "amc": amc, "fund_code": ws.title, "fund_name": fund_name, "_ident": ident,
        "period": period, "as_of": as_of, "source_url": source_url,
        "raw": raw_rel, "fetched": dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat(),
        "net_assets_cr": net, "n_holdings": len(holdings), "allocation": alloc,
        "sectors": sorted(({"sector": k, "weight": v} for k, v in sectors.items()),
                          key=lambda x: -x["weight"]),
        "derivatives": derivatives, "top": holdings[:10], "holdings": holdings,
    }


def _derivatives(rows):
    out = []
    for i in range(len(rows)):
        cells = [str(x).lower() if x is not None else "" for x in rows[i]]
        joined = " | ".join(cells)
        if "long/short" in joined and ("% to net asset" in joined or "% to nav" in joined):
            ls = next((j for j, x in enumerate(cells) if "long/short" in x), -1)
            pc = next((j for j, x in enumerate(cells) if "% to net asset" in x or "% to nav" in x), -1)
            nm = next((j for j, x in enumerate(cells) if "underlying" in x), 1)
            for r in rows[i + 1:]:
                pos = str(r[ls]).strip() if 0 <= ls < len(r) and r[ls] else ""
                if pos.lower() not in ("long", "short"):
                    break
                p = to_float(r[pc]) if 0 <= pc < len(r) else None
                n = str(r[nm]).strip() if 0 <= nm < len(r) and r[nm] else ""
                if n and p is not None:
                    out.append({"underlying": n, "position": pos.title(), "weight": round(p * 100, 4)})
    return out


def parse_sebi_xlsx(blob, source_url, meta, raw_rel):
    """Parse an xlsx (or a zip bundle of xlsx) -> list of per-fund portfolio dicts."""
    import openpyxl
    books = []
    if blob[:2] == b"PK" and b"[Content_Types].xml" not in blob[:4000] and _is_zip_bundle(blob):
        with zipfile.ZipFile(io.BytesIO(blob)) as z:
            for nm in z.namelist():
                if nm.lower().endswith((".xlsx", ".xlsm")):
                    books.append(z.read(nm))
    else:
        books = [blob]
    funds = []
    for b in books:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(b), read_only=True, data_only=True)
        except Exception as e:
            sys.stderr.write(f"  ! openpyxl: {e}\n")
            continue
        for sn in wb.sheetnames:
            if sn.lower() == "index":
                continue
            f = parse_sebi_xlsx_one(wb[sn], meta, meta["amc"], source_url, raw_rel)
            if f:
                funds.append(f)
    return funds


def _is_zip_bundle(blob):
    try:
        with zipfile.ZipFile(io.BytesIO(blob)) as z:
            names = [n.lower() for n in z.namelist()]
        return any(n.endswith((".xlsx", ".xlsm")) for n in names)
    except zipfile.BadZipFile:
        return False


ADAPTERS = {"sebi_xlsx": parse_sebi_xlsx}


# --------------------------------------------------------------------------- #
# storage
# --------------------------------------------------------------------------- #

def out_dir():
    return os.path.join(HERE, "portfolios")


def save_raw(amc, period, filename, data):
    if not filename or "." not in filename:
        filename = (filename or "portfolio") + (".zip" if _is_zip_bundle(data) else ".xlsx")
    d = os.path.join(out_dir(), "raw", amc, period)
    os.makedirs(d, exist_ok=True)
    rel = f"portfolios/raw/{amc}/{period}/{filename}"
    with open(os.path.join(HERE, rel), "wb") as f:
        f.write(data)
    return rel


def save_fund(fund):
    d = os.path.join(out_dir(), "data", fund["period"])
    os.makedirs(d, exist_ok=True)
    safe = re.sub(r"[^A-Za-z0-9_-]", "-", fund["fund_code"])[:40]
    rel = f'portfolios/data/{fund["period"]}/{fund["amc"]}__{safe}.json'
    with open(os.path.join(HERE, rel), "w", encoding="utf-8") as f:
        json.dump(fund, f, ensure_ascii=False, indent=1)
    return rel


def rebuild_index():
    reg = load_registry()
    schemes, periods = {}, set()
    amcs = {}
    for k, v in reg["amcs"].items():
        amcs[k] = {"name": v.get("name", k), "brand": v.get("brand", ""),
                   "parent": v.get("parent", ""), "site": v.get("site", ""),
                   "nav_sif": v.get("nav_sif", ""),
                   "disclosure_page": v.get("disclosure_page", v.get("site", "")),
                   "managers": v.get("managers", ""), "strategy": v.get("strategy", ""),
                   "presentation": v.get("presentation", "")}
    data_root = os.path.join(out_dir(), "data")
    if os.path.isdir(data_root):
        for period in sorted(os.listdir(data_root)):
            pdir = os.path.join(data_root, period)
            if not os.path.isdir(pdir):
                continue
            for fn in sorted(os.listdir(pdir)):
                if not fn.endswith(".json"):
                    continue
                rel = f"portfolios/data/{period}/{fn}"
                with open(os.path.join(pdir, fn), encoding="utf-8") as f:
                    fund = json.load(f)
                periods.add(period)
                for code in fund.get("_codes", []):
                    e = schemes.setdefault(code, {"name": fund["fund_name"], "amc": fund["amc"],
                                                  "fund": fund["fund_name"], "periods": {}})
                    e["periods"][period] = rel
    idx = {"generated": dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat(),
           "note": "SIFs disclose portfolios bi-monthly. Source: each AMC's own website (registry.json).",
           "periods": sorted(periods), "amcs": amcs, "schemes": schemes}
    with open(os.path.join(out_dir(), "index.json"), "w", encoding="utf-8") as f:
        json.dump(idx, f, ensure_ascii=False, indent=1)
    return idx


# --------------------------------------------------------------------------- #
# orchestration
# --------------------------------------------------------------------------- #

def _write_funds(funds, amc_schemes, cfg):
    n = 0
    overrides = cfg.get("funds", {})
    for fund in funds:
        ident = fund.pop("_ident", "")
        codes, strat, disp = map_fund(ident, fund["fund_name"], amc_schemes)
        fund["_codes"] = codes
        fund["fund_name"] = disp
        # per-fund manager/strategy/presentation, with AMC-level fallback
        o = overrides.get(strat, {}) if strat else {}
        fund["managers"] = o.get("managers", cfg.get("managers", ""))
        fund["strategy"] = o.get("strategy", cfg.get("strategy", ""))
        fund["presentation"] = o.get("presentation", cfg.get("presentation", ""))
        rel = save_fund(fund)
        n += 1
        sys.stderr.write(f"    {disp}: {fund['n_holdings']} holdings "
                         f"-> {len(codes)} code(s) -> {rel}\n")
    return n


def process_amc(amc, cfg, period, nav_schemes):
    disc = cfg.get("discovery", {})
    method = disc.get("method")
    discover = DISCOVERERS.get(method)
    parse = ADAPTERS.get(cfg.get("adapter", "sebi_xlsx"))
    if not discover or not parse:
        sys.stderr.write(f"  ! {amc}: no discovery '{method}' / adapter\n")
        return 0
    try:
        files = discover(cfg)
    except Exception as e:
        sys.stderr.write(f"  ! {amc}: discovery failed: {e}\n")
        return 0
    if not files:
        sys.stderr.write(f"  · {amc}: no disclosure files found (may not have disclosed yet)\n")
        return 0
    target = period or files[-1]["period"]         # latest disclosed period
    files = [f for f in files if f["period"] == target]   # ALL files for that period
    if not files:
        sys.stderr.write(f"  · {amc}: nothing for period {period}\n")
        return 0

    amc_schemes = [s for s in nav_schemes if norm(s["sif"]) == norm(cfg.get("nav_sif", ""))]
    written = 0
    for fmeta in files:
        blob = fmeta.get("_blob")
        if blob is None:
            try:
                blob = http_get(fmeta["url"])
            except Exception as e:
                sys.stderr.write(f"  ! {amc} {fmeta['url']}: {e}\n")
                continue
        if not looks_like_office(blob) and blob[:4] != b"%PDF":
            sys.stderr.write(f"  ! {amc}: {fmeta['url']} did not return a file (likely WAF/HTML). "
                             f"Use a browser or pin a manual URL.\n")
            continue
        raw_rel = save_raw(amc, fmeta["period"], fmeta["filename"], blob)
        meta = {"amc": amc, "period": fmeta["period"], "as_of": fmeta["as_of"],
                "raw": raw_rel, "filename": fmeta["filename"]}
        sys.stderr.write(f"[{amc}] {fmeta['period']} ({fmeta['filename']})\n")
        funds = parse(blob, fmeta["url"], meta, raw_rel)
        written += _write_funds(funds, amc_schemes, cfg)
    return written


def process_file(path, amc, period, nav_schemes):
    with open(path, "rb") as f:
        blob = f.read()
    reg = load_registry()
    cfg = reg["amcs"].get(amc, {})
    as_of = f"{period}-28" if period else dt.date.today().isoformat()
    per = period or dt.date.today().strftime("%Y-%m")
    raw_rel = save_raw(amc, per, os.path.basename(path), blob)
    meta = {"amc": amc, "period": per, "as_of": as_of, "raw": raw_rel,
            "filename": os.path.basename(path)}
    parse = ADAPTERS.get(cfg.get("adapter", "sebi_xlsx"), parse_sebi_xlsx)
    amc_schemes = [s for s in nav_schemes if norm(s["sif"]) == norm(cfg.get("nav_sif", ""))]
    funds = parse(blob, "file://" + os.path.basename(path), meta, raw_rel)
    sys.stderr.write(f"[{amc}] {per} ({os.path.basename(path)})\n")
    return _write_funds(funds, amc_schemes, cfg)


def main(argv=None):
    p = argparse.ArgumentParser(description="Fetch & normalise SIF portfolio disclosures.")
    p.add_argument("--amc", help="AMC key from registry.json (default: all)")
    p.add_argument("--period", help="YYYY-MM to fetch (default: latest disclosed)")
    p.add_argument("--file", help="normalise a hand-downloaded xlsx/zip instead of fetching")
    p.add_argument("--list", action="store_true", help="list stored portfolios and exit")
    args = p.parse_args(argv)

    if args.list:
        idx = rebuild_index()
        print(f"periods: {', '.join(idx['periods']) or '(none)'}")
        print(f"scheme codes with disclosures: {len(idx['schemes'])}")
        return 0

    nav_schemes = load_nav_schemes()
    if args.file:
        if not args.amc:
            p.error("--file requires --amc")
        process_file(args.file, args.amc, args.period, nav_schemes)
        rebuild_index()
        return 0

    reg = load_registry()
    amcs = [args.amc] if args.amc else list(reg["amcs"].keys())
    total = 0
    for amc in amcs:
        cfg = reg["amcs"].get(amc)
        if not cfg:
            sys.stderr.write(f"  ! unknown AMC '{amc}'\n")
            continue
        total += process_amc(amc, cfg, args.period, nav_schemes)
    idx = rebuild_index()
    sys.stderr.write(f"[done] {total} fund portfolio(s) written; "
                     f"{len(idx['schemes'])} scheme code(s) now have disclosures.\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
